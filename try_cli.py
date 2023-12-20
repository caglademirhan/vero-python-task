import argparse
import csv
import json
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Authorization details
BASE_URL = "https://api.baubuddy.de/dev/index.php/v1"
LOGIN_URL = "https://api.baubuddy.de/index.php/login"
USERNAME = "365"
PASSWORD = "1"
HEADERS = {
    "Authorization": "Basic QVBJX0V4cGxvcmVyOjEyMzQ1NmlzQUxhbWVQYXNz",
    "Content-Type": "application/json",
}

def get_access_token():
    payload = {"username": USERNAME, "password": PASSWORD}
    response = requests.post(LOGIN_URL, json=payload, headers=HEADERS)
    access_token = response.json()["oauth"]["access_token"]
    return access_token

def get_csv_data(file):
    """Extracts list of dictionaries from csv file."""
    csv_dicts = []
    with open(file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for line in reader:
            csv_dicts.append(line)
    return csv_dicts

def get_resources(csv_data, keys, colored):
    access_token = get_access_token()
    url = f"{BASE_URL}/vehicles/select/active"
    headers = {"Authorization": f"Bearer {access_token}"}
    payload = {"keys": keys, "colored": colored}

    # Read the contents of the CSV file
    with open("vehicles.csv", "r", encoding="utf-8-sig") as file:
        csv_content = file.read()

    files = {"file": ("vehicles.csv", csv_content, "text/csv")}  # Specify the file type
    
    try:
        response = requests.post(url, data=payload, files=files, headers=headers)
        response.raise_for_status()  # Raise an HTTPError for bad responses (4xx or 5xx)
        
        try:
            json_data = response.json()
            return json_data
        except ValueError as ve:
            print(f"Error decoding JSON: {ve}")
            print("Response content:", response.content)
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error in get_resources: {e}")
        return None  # Return None if an error occurs


def create_excel(data, keys, colored):
    # Sorting rows by 'gruppe' if it exists, otherwise use an empty string
    data = sorted(data, key=lambda x: x.get('gruppe', ''))

    # Create a new workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['rnr'] + keys
    worksheet.append(headers)

    # Apply styles to headers
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Fill the worksheet with data
    for row_data in data:
        row = [row_data.get(header, '') for header in headers]
        worksheet.append(row)

    # Apply coloring logic if -c flag is True
    if colored:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
            cell_value = row[0].value
            hu_date = datetime.strptime(cell_value, '%Y-%m-%d')
            now_date = datetime.now()
            diff_in_months = (now_date - hu_date) // timedelta(days=30)

            if diff_in_months < 3:
                fill_color = '00FF00'  # Green
            elif diff_in_months <= 12:
                fill_color = 'FFA500'  # Orange
            else:
                fill_color = 'B30000'  # Red

            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Save the workbook
    current_date_iso_formatted = datetime.now().isoformat().replace(":", "_")
    excel_filename = f"vehicles_{current_date_iso_formatted}.xlsx"
    workbook.save(excel_filename)

def main():
    parser = argparse.ArgumentParser(description="Client script for Baubuddy API")
    parser.add_argument("-k", "--keys", nargs="+", help="List of keys to consider")
    parser.add_argument("-c", "--colored", action="store_true", help="Color each row based on logic")
    args = parser.parse_args()

    # Get data from CSV file
    csv_data = get_csv_data("vehicles.csv")

    # Get resources from the server
    response_data = get_resources(csv_data, args.keys, args.colored)

    if response_data is not None:
        # Convert the server's response to an Excel file
        create_excel(response_data, args.keys, args.colored)

if __name__ == "__main__":
    main()





